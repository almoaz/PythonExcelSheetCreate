from openpyxl import Workbook  # importing our library
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

import pyrebase
#auth.uid != null

const  = {
  "apiKey": "AIzaSyBLOfT0N-JTkm79S-XKvTIMvw3gNBBLsis",
  "authDomain": "smuct-6c3ad.firebaseapp.com",
  "databaseURL": "https://smuct-6c3ad-default-rtdb.firebaseio.com/",
  "projectId": "smuct-6c3ad",
  "storageBucket": "smuct-6c3ad.appspot.com",
  "messagingSenderId": "1013082282201",
  "appId": "1:1013082282201:web:90172c1a8806071c23c6c8",
  "measurementId": "G-BP3NWNMGT1"};

firebase = pyrebase.initialize_app(const)
db = firebase.database()

import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showinfo

# root window
root = tk.Tk()
root.geometry("350x250")
root.resizable(False, False)
root.title('Routine Excel Sheet Generate')

# store email address and password
email = tk.StringVar()
password = tk.StringVar()


def login_clicked():
    """ callback when the login button clicked
    """
    wb = Workbook()
    sheet = wb.active

    # marge cell
    sheet.merge_cells('A1:N4')

    sheet.merge_cells('C5:F5')
    sheet.merge_cells('G5:N5')

    sheet.merge_cells('A6:A7')
    sheet.merge_cells('B6:B7')

    sheet.merge_cells('C6:D6')
    sheet.merge_cells('E6:F6')
    sheet.merge_cells('G6:H6')
    sheet.merge_cells('I6:J6')
    sheet.merge_cells('K6:L6')
    sheet.merge_cells('M6:N6')

    sheet.merge_cells('A8:A19')
    sheet.merge_cells('A20:A31')
    sheet.merge_cells('A32:A43')
    sheet.merge_cells('A44:A55')
    sheet.merge_cells('A56:A67')
    sheet.merge_cells('A68:A79')
    sheet.merge_cells('A80:A84')

    sheet.merge_cells('A86:N97')

    # for border

    tblr = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))

    for x in range(8):

        if x == 5:

            for y in range(14):
                cell = sheet.cell(row=x, column=1 + y)
                cell.fill = PatternFill("solid", start_color="ffb366")

        if x > 5:

            for y in range(14):
                cell = sheet.cell(row=x, column=1 + y)
                cell.fill = PatternFill("solid", start_color="99ff99")

        print(x)

    for x in range(84):

        if x > 7:
            cell = sheet.cell(row=x, column=1)
            cell.fill = PatternFill("solid", start_color="ffb366")

            if x == 8:
                cell = sheet.cell(row=x, column=1)
                cell.value = "Satureday"
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if x == 20:
                cell = sheet.cell(row=x, column=1)
                cell.value = "Sunday"
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if x == 32:
                cell = sheet.cell(row=x, column=1)
                cell.value = "Monday"
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if x == 44:
                cell = sheet.cell(row=x, column=1)
                cell.value = "Tuesday"
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if x == 56:
                cell = sheet.cell(row=x, column=1)
                cell.value = "Wednesday"
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if x == 68:
                cell = sheet.cell(row=x, column=1)
                cell.value = "Thursday"
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if x == 80:
                cell = sheet.cell(row=x, column=1)
                cell.value = "Friday"
                cell.alignment = Alignment(horizontal='center', vertical='center')

    for x in range(98):

        if x > 0:

            for y in range(14):
                cell = sheet.cell(row=x, column=1 + y)
                cell.border = tblr

        print(x)

    for x in range(15):
        if x == 3:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Module"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 4:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Faculty"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 5:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Module"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 6:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Faculty"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 7:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Module"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 8:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Faculty"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 9:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Module"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 10:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Faculty"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 11:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Module"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 12:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Faculty"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 13:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Module"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 14:
            cell = sheet.cell(row=7, column=x)
            cell.value = "Faculty"
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for x in range(84):

        if x == 1:
            cell = sheet.cell(row=7 + x, column=2)
            cell.value = "Tri-1st"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=8 + x, column=2)
            cell.value = "Tri-2nd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=9 + x, column=2)
            cell.value = "Tri-3rd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=10 + x, column=2)
            cell.value = "Tri-4th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=11 + x, column=2)
            cell.value = "Tri-5th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=12 + x, column=2)
            cell.value = "Tri-6th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=13 + x, column=2)
            cell.value = "Tri-7th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=14 + x, column=2)
            cell.value = "Tri-8th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=15 + x, column=2)
            cell.value = "Tri-9th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=16 + x, column=2)
            cell.value = "Tri-10th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=17 + x, column=2)
            cell.value = "Tri-11th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=18 + x, column=2)
            cell.value = "Tri-12th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 1:
            cell = sheet.cell(row=19 + x, column=2)
            cell.value = "Tri-1st"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=20 + x, column=2)
            cell.value = "Tri-2nd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=21 + x, column=2)
            cell.value = "Tri-3rd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=22 + x, column=2)
            cell.value = "Tri-4th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=23 + x, column=2)
            cell.value = "Tri-5th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=24 + x, column=2)
            cell.value = "Tri-6th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=25 + x, column=2)
            cell.value = "Tri-7th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=26 + x, column=2)
            cell.value = "Tri-8th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=27 + x, column=2)
            cell.value = "Tri-9th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=28 + x, column=2)
            cell.value = "Tri-10th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=29 + x, column=2)
            cell.value = "Tri-11th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=30 + x, column=2)
            cell.value = "Tri-12th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 1:
            cell = sheet.cell(row=31 + x, column=2)
            cell.value = "Tri-1st"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=32 + x, column=2)
            cell.value = "Tri-2nd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=33 + x, column=2)
            cell.value = "Tri-3rd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=34 + x, column=2)
            cell.value = "Tri-4th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=35 + x, column=2)
            cell.value = "Tri-5th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=36 + x, column=2)
            cell.value = "Tri-6th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=37 + x, column=2)
            cell.value = "Tri-7th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=38 + x, column=2)
            cell.value = "Tri-8th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=39 + x, column=2)
            cell.value = "Tri-9th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=40 + x, column=2)
            cell.value = "Tri-10th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=41 + x, column=2)
            cell.value = "Tri-11th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=42 + x, column=2)
            cell.value = "Tri-12th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 1:
            cell = sheet.cell(row=43 + x, column=2)
            cell.value = "Tri-1st"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=44 + x, column=2)
            cell.value = "Tri-2nd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=45 + x, column=2)
            cell.value = "Tri-3rd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=46 + x, column=2)
            cell.value = "Tri-4th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=47 + x, column=2)
            cell.value = "Tri-5th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=48 + x, column=2)
            cell.value = "Tri-6th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=49 + x, column=2)
            cell.value = "Tri-7th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=50 + x, column=2)
            cell.value = "Tri-8th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=51 + x, column=2)
            cell.value = "Tri-9th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=52 + x, column=2)
            cell.value = "Tri-10th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=53 + x, column=2)
            cell.value = "Tri-11th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=54 + x, column=2)
            cell.value = "Tri-12th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 1:
            cell = sheet.cell(row=55 + x, column=2)
            cell.value = "Tri-1st"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=56 + x, column=2)
            cell.value = "Tri-2nd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=57 + x, column=2)
            cell.value = "Tri-3rd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=58 + x, column=2)
            cell.value = "Tri-4th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=59 + x, column=2)
            cell.value = "Tri-5th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=60 + x, column=2)
            cell.value = "Tri-6th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=61 + x, column=2)
            cell.value = "Tri-7th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=62 + x, column=2)
            cell.value = "Tri-8th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=63 + x, column=2)
            cell.value = "Tri-9th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=64 + x, column=2)
            cell.value = "Tri-10th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=65 + x, column=2)
            cell.value = "Tri-11th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=66 + x, column=2)
            cell.value = "Tri-12th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if x == 1:
            cell = sheet.cell(row=67 + x, column=2)
            cell.value = "Tri-1st"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=68 + x, column=2)
            cell.value = "Tri-2nd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=69 + x, column=2)
            cell.value = "Tri-3rd"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=70 + x, column=2)
            cell.value = "Tri-4th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=71 + x, column=2)
            cell.value = "Tri-5th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=72 + x, column=2)
            cell.value = "Tri-6th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=73 + x, column=2)
            cell.value = "Tri-7th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=74 + x, column=2)
            cell.value = "Tri-8th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=75 + x, column=2)
            cell.value = "Tri-9th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=76 + x, column=2)
            cell.value = "Tri-10th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=77 + x, column=2)
            cell.value = "Tri-11th"
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=78 + x, column=2)
            cell.value = "Tri-12th"
            cell.alignment = Alignment(horizontal='center', vertical='center')
    txt = 'aabcdefgh'

    for y in range(7):
        t = txt[y]
        if y > 0 and y < 7:
            course = db.child("Department Of CSE And CSIT").child("Routine").child("Time").child(t+"Time").get()
            # print(t + "1C: " + course.val() + t + "1T: " + teacher.val() + t + "1R: " + Room.val())
            cell = sheet.cell(row=6, column=1 + y + y)
            cell.value = course.val()
            cell.alignment = Alignment(horizontal='center', vertical='center')

    text = 'AABCDEFABCDEFABCDEFABCDEF'

    for y in range(12):

        if y > 0:
            for x in range(7):
                t = text[x]

                if x > 0 and x < 7:
                    course = db.child("Department Of CSE And CSIT").child("Routine").child("Saturday").child(
                        t + str(y) + "C").get()
                    teacher = db.child("Department Of CSE And CSIT").child("Routine").child("Saturday").child(
                        t + str(y) + "T").get()
                    Room = db.child("Department Of CSE And CSIT").child("Routine").child("Saturday").child(
                        t + str(y) + "R").get()
                    # print(t + "1C: " + course.val() + t + "1T: " + teacher.val() + t + "1R: " + Room.val())
                    cell = sheet.cell(row=7 + y, column=1 + x + x)
                    cell.value = course.val()
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    cell = sheet.cell(row=7 + y, column=2 + x + x)
                    if teacher.val() != '' or Room.val() != '':
                        cell.value = teacher.val() + ' (' + Room.val() + '505' + ')'
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    for y in range(12):

        if y > 0:
            for x in range(7):
                t = text[x]

                if x > 0 and x < 7:
                    course = db.child("Department Of CSE And CSIT").child("Routine").child("Sunday").child(
                        t + str(y) + "C").get()
                    teacher = db.child("Department Of CSE And CSIT").child("Routine").child("Sunday").child(
                        t + str(y) + "T").get()
                    Room = db.child("Department Of CSE And CSIT").child("Routine").child("Sunday").child(
                        t + str(y) + "R").get()
                    # print(t + "1C: " + course.val() + t + "1T: " + teacher.val() + t + "1R: " + Room.val())
                    cell = sheet.cell(row=19 + y, column=1 + x + x)
                    cell.value = course.val()
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    cell = sheet.cell(row=19 + y, column=2 + x + x)
                    if teacher.val() != '' or Room.val() != '':
                        cell.value = teacher.val() + ' (' + Room.val() + '505' + ')'
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    for y in range(12):

        if y > 0:
            for x in range(7):
                t = text[x]

                if x > 0 and x < 7:
                    course = db.child("Department Of CSE And CSIT").child("Routine").child("Monday").child(
                        t + str(y) + "C").get()
                    teacher = db.child("Department Of CSE And CSIT").child("Routine").child("Monday").child(
                        t + str(y) + "T").get()
                    Room = db.child("Department Of CSE And CSIT").child("Routine").child("Monday").child(
                        t + str(y) + "R").get()
                    # print(t + "1C: " + course.val() + t + "1T: " + teacher.val() + t + "1R: " + Room.val())
                    cell = sheet.cell(row=31 + y, column=1 + x + x)
                    cell.value = course.val()
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    cell = sheet.cell(row=31 + y, column=2 + x + x)
                    if teacher.val() != '' or Room.val() != '':
                        cell.value = teacher.val() + ' (' + Room.val() + '505' + ')'
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    for y in range(12):

        if y > 0:
            for x in range(7):
                t = text[x]

                if x > 0 and x < 7:
                    course = db.child("Department Of CSE And CSIT").child("Routine").child("Tuesday").child(
                        t + str(y) + "C").get()
                    teacher = db.child("Department Of CSE And CSIT").child("Routine").child("Tuesday").child(
                        t + str(y) + "T").get()
                    Room = db.child("Department Of CSE And CSIT").child("Routine").child("Tuesday").child(
                        t + str(y) + "R").get()
                    # print(t + "1C: " + course.val() + t + "1T: " + teacher.val() + t + "1R: " + Room.val())
                    cell = sheet.cell(row=43 + y, column=1 + x + x)
                    cell.value = course.val()
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    cell = sheet.cell(row=43 + y, column=2 + x + x)
                    if teacher.val() != '' or Room.val() != '':
                        cell.value = teacher.val() + ' (' + Room.val() + '505' + ')'
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    for y in range(12):

        if y > 0:
            for x in range(7):
                t = text[x]

                if x > 0 and x < 7:
                    course = db.child("Department Of CSE And CSIT").child("Routine").child("Wednesday").child(
                        t + str(y) + "C").get()
                    teacher = db.child("Department Of CSE And CSIT").child("Routine").child("Wednesday").child(
                        t + str(y) + "T").get()
                    Room = db.child("Department Of CSE And CSIT").child("Routine").child("Wednesday").child(
                        t + str(y) + "R").get()
                    # print(t + "1C: " + course.val() + t + "1T: " + teacher.val() + t + "1R: " + Room.val())
                    cell = sheet.cell(row=55 + y, column=1 + x + x)
                    cell.value = course.val()
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    cell = sheet.cell(row=55 + y, column=2 + x + x)
                    if teacher.val() != '' or Room.val() != '':
                        cell.value = teacher.val() + ' (' + Room.val() + '505' + ')'
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    for y in range(12):

        if y > 0:
            for x in range(7):
                t = text[x]

                if x > 0 and x < 7:
                    course = db.child("Department Of CSE And CSIT").child("Routine").child("Thursday").child(
                        t + str(y) + "C").get()
                    teacher = db.child("Department Of CSE And CSIT").child("Routine").child("Thursday").child(
                        t + str(y) + "T").get()
                    Room = db.child("Department Of CSE And CSIT").child("Routine").child("Thursday").child(
                        t + str(y) + "R").get()
                    # print(t + "1C: " + course.val() + t + "1T: " + teacher.val() + t + "1R: " + Room.val())
                    cell = sheet.cell(row=67 + y, column=1 + x + x)
                    cell.value = course.val()
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    cell = sheet.cell(row=67 + y, column=2 + x + x)
                    if teacher.val() != '' or Room.val() != '':
                        cell.value = teacher.val() + ' (' + Room.val() + '505' + ')'
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=1, column=1)
    cell.value = "SHANTO-MARIAM UNIVERSITY OF CREATIVE TECHNOLOGY"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=5, column=3)
    cell.value = "Morning"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=5, column=7)
    cell.value = "Evening"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=6, column=1)
    cell.value = "Day"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=6, column=2)
    cell.value = "Semester"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save('C:\\Users\Mahfuz Salehin Moaz\Desktop\\routine_cse_smuct.xlsx')

    print("Successfully routine excel sheet generate")




# login button
login_button = ttk.Button(text="Generate", command=login_clicked)
login_button.pack(fill='x', expand=True, pady=10)


root.mainloop()



























'''cell = sheet.cell(row=1, column=1)

cell.value = "SHANTO-MARIAM UNIVERSITY OF CREATIVE TECHNOLOGY" \
             "   Departments of CSE and CSIT (Day Program)" \
             "          Trimester Spring 2022"
cell.alignment = Alignment(horizontal='center', vertical='center')




cell2 = sheet.cell(row=5, column=3)

cell2.fill = PatternFill("solid", start_color="ffb366")
cell2.value = "Morning"
cell2.alignment = Alignment(horizontal='center', vertical='center')



#cell2.border = thin_border

cell3 = sheet.cell(row=5, column=7)
cell3.fill = PatternFill("solid", start_color="ffb366")
cell3.value = "Evening"
cell3.alignment = Alignment(horizontal='center', vertical='center')

cell4 = sheet.cell(row=5, column=8)
cell5 = sheet.cell(row=5, column=9)
cell6 = sheet.cell(row=5, column=10)
cell7 = sheet.cell(row=5, column=11)
cell8 = sheet.cell(row=5, column=12)
cell9 = sheet.cell(row=5, column=13)
cell10 = sheet.cell(row=5, column=14)'''





