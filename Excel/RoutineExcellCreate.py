from openpyxl import Workbook  # importing our library
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = Workbook()
sheet = wb.active


sheet.merge_cells('A1:B1')
sheet.merge_cells('A2:B3')
sheet.merge_cells('A4:B5')
sheet.merge_cells('A6:B7')
sheet.merge_cells('A8:B9')
sheet.merge_cells('A10:B11')
sheet.merge_cells('A12:B13')
sheet.merge_cells('A14:B15')
sheet.merge_cells('A16:B17')
sheet.merge_cells('A18:B19')
sheet.merge_cells('A20:B21')
sheet.merge_cells('A22:B23')
sheet.merge_cells('A24:B25')

cell = sheet.cell(row=1, column=1)
cell.value = 'All Semester'
cell.alignment = Alignment(horizontal='center', vertical='center')

cell1 = sheet.cell(row=2, column=1)
cell1.value = '1st Semester'
cell1.alignment = Alignment(horizontal='center', vertical='center')

cell2 = sheet.cell(row=4, column=1)
cell2.value = '2nd Semester'
cell2.alignment = Alignment(horizontal='center', vertical='center')

cell3 = sheet.cell(row=6, column=1)
cell3.value = '3rd Semester'
cell3.alignment = Alignment(horizontal='center', vertical='center')

cell4 = sheet.cell(row=8, column=1)
cell4.value = '4th Semester'
cell4.alignment = Alignment(horizontal='center', vertical='center')

cell5 = sheet.cell(row=10, column=1)
cell5.value = '5th Semester'
cell5.alignment = Alignment(horizontal='center', vertical='center')

cell6 = sheet.cell(row=12, column=1)
cell6.value = '6th Semester'
cell6.alignment = Alignment(horizontal='center', vertical='center')

cell7 = sheet.cell(row=14, column=1)
cell7.value = '7th Semester'
cell7.alignment = Alignment(horizontal='center', vertical='center')

cell8 = sheet.cell(row=16, column=1)
cell8.value = '8th Semester'
cell8.alignment = Alignment(horizontal='center', vertical='center')

cell9 = sheet.cell(row=18, column=1)
cell9.value = '9th Semester'
cell9.alignment = Alignment(horizontal='center', vertical='center')

cell10 = sheet.cell(row=20, column=1)
cell10.value = '10th Semester'
cell10.alignment = Alignment(horizontal='center', vertical='center')

cell11 = sheet.cell(row=22, column=1)
cell11.value = '11th Semester'
cell11.alignment = Alignment(horizontal='center', vertical='center')

cell12 = sheet.cell(row=24, column=1)
cell12.value = '12th Semester'
cell12.alignment = Alignment(horizontal='center', vertical='center')

wb.save('C:\\Users\Mahfuz Salehin Moaz\Desktop\\routine.xlsx')


